"""
export_project_data_to_excel.py — PY1 (10 pts)
Exports flower shop database to a timestamped Excel file.
Requires: pip install pymysql pandas openpyxl
"""

import os
import sys
import traceback
from datetime import datetime

try:
    import pymysql
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"[ERROR] Missing library: {e}")
    print("Run:  pip install pymysql pandas openpyxl")
    sys.exit(1)

# ── DB Configuration ────────────────────────────────────────────
DB_CONFIG = {
    "host":     "localhost",
    "port":     3306,
    "user":     "root",
    "password": "",           # Default XAMPP password
    "db":       "flower_shop_dss",
    "charset":  "utf8mb4",
}

# ── Output folder ────────────────────────────────────────────────
EXPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
filename  = f"flower_shop_export_{timestamp}.xlsx"
filepath  = os.path.join(EXPORT_DIR, filename)

print(f"Petals & Bloom — Excel Export")
print(f"Connecting to MySQL ({DB_CONFIG['host']}:{DB_CONFIG['port']})…")

try:
    conn = pymysql.connect(**DB_CONFIG)
    print("Connected.")
except pymysql.Error as e:
    print(f"[ERROR] DB connection failed: {e}")
    sys.exit(1)

try:
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:

        # ── Sheet 1: Customers ──────────────────────────────────
        df_cust = pd.read_sql(
            "SELECT id, CONCAT(first_name,' ',last_name) AS full_name, email, phone, address, created_at FROM customers ORDER BY created_at DESC",
            conn
        )
        df_cust.to_excel(writer, sheet_name="Customers", index=False)
        print(f"  Sheet 'Customers': {len(df_cust)} rows")

        # ── Sheet 2: Orders ─────────────────────────────────────
        df_orders = pd.read_sql(
            """SELECT o.id, o.order_code, o.customer_name, o.customer_email, o.customer_phone,
                      o.delivery_address, o.delivery_date, o.occasion, o.status,
                      o.payment_method, o.payment_status, o.total_price,
                      o.card_message, o.special_notes, o.created_at
               FROM orders o ORDER BY o.created_at DESC""",
            conn
        )
        df_orders["delivery_date"] = pd.to_datetime(df_orders["delivery_date"]).dt.strftime("%Y-%m-%d")
        df_orders["created_at"]    = pd.to_datetime(df_orders["created_at"]).dt.strftime("%Y-%m-%d %H:%M")
        df_orders.to_excel(writer, sheet_name="Orders", index=False)
        print(f"  Sheet 'Orders': {len(df_orders)} rows")

        # ── Sheet 3: Order Items ─────────────────────────────────
        df_items = pd.read_sql(
            """SELECT oi.id, oi.order_id, o.order_code, p.name AS product_name,
                      oi.quantity, oi.unit_price, oi.subtotal
               FROM order_items oi
               JOIN orders o  ON oi.order_id  = o.id
               JOIN products p ON oi.product_id = p.id
               ORDER BY oi.order_id""",
            conn
        )
        df_items.to_excel(writer, sheet_name="Order Items", index=False)
        print(f"  Sheet 'Order Items': {len(df_items)} rows")

        # ── Sheet 4: Products ───────────────────────────────────
        df_prods = pd.read_sql(
            """SELECT p.id, p.name, c.name AS category, p.price,
                      p.stock_quantity, p.is_featured, p.is_active, p.created_at
               FROM products p LEFT JOIN categories c ON p.category_id = c.id
               ORDER BY p.name""",
            conn
        )
        df_prods.to_excel(writer, sheet_name="Products", index=False)
        print(f"  Sheet 'Products': {len(df_prods)} rows")

        # ── Sheet 5: KPI Summary ─────────────────────────────────
        kpi_data = {}

        r = pd.read_sql("SELECT COUNT(*) AS n FROM orders", conn).iloc[0]
        kpi_data["Total Orders"] = int(r["n"])

        r = pd.read_sql("SELECT COUNT(*) AS n FROM orders WHERE status='delivered'", conn).iloc[0]
        kpi_data["Delivered Orders"] = int(r["n"])

        r = pd.read_sql("SELECT COUNT(*) AS n FROM orders WHERE status='cancelled'", conn).iloc[0]
        kpi_data["Cancelled Orders"] = int(r["n"])

        r = pd.read_sql("SELECT COALESCE(SUM(total_price),0) AS s FROM orders WHERE payment_status='paid'", conn).iloc[0]
        kpi_data["Total Revenue (RON)"] = float(r["s"])

        r = pd.read_sql("SELECT COALESCE(AVG(total_price),0) AS a FROM orders", conn).iloc[0]
        kpi_data["Average Order Value (RON)"] = round(float(r["a"]), 2)

        r = pd.read_sql("SELECT COUNT(*) AS n FROM customers", conn).iloc[0]
        kpi_data["Total Customers"] = int(r["n"])

        r = pd.read_sql("SELECT COUNT(*) AS n FROM products WHERE is_active=1", conn).iloc[0]
        kpi_data["Active Products"] = int(r["n"])

        r = pd.read_sql("SELECT COUNT(*) AS n FROM products WHERE stock_quantity<=5 AND is_active=1", conn).iloc[0]
        kpi_data["Low Stock Products (≤5)"] = int(r["n"])

        # Top category by orders
        r = pd.read_sql(
            """SELECT c.name, COUNT(oi.id) AS cnt FROM order_items oi
               JOIN products p ON oi.product_id=p.id
               JOIN categories c ON p.category_id=c.id
               GROUP BY c.id ORDER BY cnt DESC LIMIT 1""",
            conn
        )
        kpi_data["Best Category"] = r.iloc[0]["name"] if len(r) > 0 else "N/A"

        kpi_data["Export Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        df_kpi = pd.DataFrame(list(kpi_data.items()), columns=["Metric", "Value"])
        df_kpi.to_excel(writer, sheet_name="KPI Summary", index=False)
        print(f"  Sheet 'KPI Summary': {len(df_kpi)} metrics")

    # ── Apply styling ────────────────────────────────────────────
    print("Applying Excel styling…")
    wb = load_workbook(filepath)

    HEADER_FILL  = PatternFill("solid", fgColor="C2185B")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    ALT_ROW_FILL = PatternFill("solid", fgColor="FCE4EC")
    THIN_BORDER  = Border(
        bottom=Side(style="thin", color="DDDDDD")
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Header row styling
        for cell in ws[1]:
            cell.font    = HEADER_FONT
            cell.fill    = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 22

        # Auto column width + alternating row colours + freeze header
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row_idx, cell in enumerate(col, 1):
                if row_idx > 1:
                    # Alternating row fill
                    if row_idx % 2 == 0:
                        cell.fill = ALT_ROW_FILL
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(vertical="center")
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes    = "A2"

    wb.save(filepath)
    print(f"\n✓ Export complete: {filepath}")

except Exception:
    print("\n[ERROR] Export failed:")
    traceback.print_exc()
    sys.exit(1)
finally:
    conn.close()
